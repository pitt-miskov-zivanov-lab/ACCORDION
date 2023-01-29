"""
@author: Yasmine
"""

import pandas as pd
import re
import os
import argparse
import openpyxl
import networkx as nx
import pickle
import time
import logging
from markovCluster import runMarkovCluster

# define regex for valid characters in variable names
_VALID_CHARS = r'a-zA-Z0-9\@\_\/'

# This function and the following function are inherited from DySE framework
def get_model(model_file):
	"""
	This function reads the baseline model of BioRECIPES format and returns two useful dictionaries

	Parameters
	----------
	model_file : str
		The path of the baseline model file

	Returns
	-------
	model_dict : dict
		Dictionary that holds critical information of each baseline model element
	regulators : dict
		 Contains baseline model elements and corresponding regulator elements
	"""

	global _VALID_CHARS

	regulators = dict()
	model_dict = dict()


	# Load the input file containing elements and regulators
	df_model = pd.read_excel(model_file, na_values='NaN', keep_default_na = False)
	# check model format
	if df_model.columns[0].lower() == 'element attributes':
		df_model = df_model.reset_index()
		df_model = df_model.rename(columns=df_model.iloc[1]).drop([0,1]).set_index('#')

	input_col_name = [x.strip() for x in df_model.columns if ('element name' in x.lower())]
	input_col_ids = [x.strip() for x in df_model.columns if ('ids' in x.lower())]

	input_col_type = [x.strip() for x in df_model.columns if ('element type' in x.lower())]
	input_col_X = [x.strip() for x in df_model.columns if ('variable' in x.lower())]
	input_col_A = [x.strip() for x in df_model.columns if ('positive' in x.lower())]
	input_col_I = [x.strip() for x in df_model.columns if ('negative' in x.lower())]

	# set index to variable name column
	# remove empty variable names
	# append cols with the sets of regulators using .apply
	for curr_row in df_model.index:
		element_name = df_model.loc[curr_row,input_col_name[-1]].strip()
		ids = df_model.loc[curr_row,input_col_ids[0]].strip().upper().split(',')
		#print(ids)
		element_type = df_model.loc[curr_row,input_col_type[0]].strip()
		var_name = df_model.loc[curr_row,input_col_X[0]].strip()
		pos_regulators = df_model.loc[curr_row,input_col_A[0]].strip()
		neg_regulators = df_model.loc[curr_row,input_col_I[0]].strip()

		if var_name == '':
			continue

		curr = []

		if pos_regulators != '':
			curr += re.findall('['+_VALID_CHARS+']+',pos_regulators)

		if neg_regulators != '':
			curr += re.findall('['+_VALID_CHARS+']+',neg_regulators)

		# returning regulators separately for compatibility with runMarkovCluster
		regulators[var_name] = set(curr)
		model_dict[var_name] = {
			'name' : element_name,
			'ids' : ids,
			'type' : element_type,
			'regulators' : set(curr)}

	return model_dict, regulators

def getVariableName(model_dict, curr_map, ext_element_info):
	"""
	A utility function for parseExtension(), which matches the element name from the extracted event to an element in the baseline model

	Parameters
	----------
	model_dict : dict
		Dictionary that holds critical information of each baseline model element
	curr_map: dict
		Temporary dictionary that contains already matched pairs
	ext_element_info: list
		List of information for certain element in the extracted event, starting with element name

	Returns
	-------
	match : str
		The most likely matched element name in model_dict, to the element represented by ext_element_info; Otherwise, return the extended element name suffix by "_ext"
	"""

	global _VALID_CHARS

	ext_element_name = ext_element_info[0]

	# Check for valid element name
	if ext_element_name=='':
		#logging.warn('Missing element name in extensions')
		return ''
	elif re.search('[^'+_VALID_CHARS+']+',ext_element_name):
		#logging.warn(('Skipping due to invalid characters in variable name: %s') % str(ext_element_name))
		return ''

	ext_element_id = ext_element_info[3]
	ext_element_type = ext_element_info[5]

	if ext_element_name in curr_map:
		return curr_map[ext_element_name]

	# from the location and type
	match = ext_element_name + '_ext'
	confidence = 0.0
	# Iterate all names in the dictionary and find the most likely match
	for key,value in model_dict.items():

		curr_conf = 0.0
		if ext_element_id.upper() in value['ids']:
			curr_conf = 1
		elif ext_element_name.upper().startswith(value['name'].upper()) \
			or value['name'].upper().startswith(ext_element_name.upper()):
			curr_conf = 0.8

		if curr_conf>0 and value['type'].lower().startswith(ext_element_type):
			curr_conf += 1

		if curr_conf > confidence:
			match = key
			confidence = curr_conf
			if curr_conf==2: break

	curr_map[ext_element_name] = match
	return match

def parseExtension(model_dict, ext_file):
	"""
     This function parses the interactions within the reading output file (extracted events) into a set object, each component in the set has compatible format for BioRECIPES.

     Parameters
     ----------
     model_dict : dict
 		Dictionary that holds critical information of each baseline model element
     ext_file : str
         The path of the reading output file (extracted events)

     Returns
     -------
     ext_edges : set
         Each interaction within the reading output file will have the form:
         (regulator element, regulated element, type of interaction (+/-))
	"""
	regulated_col = 8
	interaction_col = 16
	ext_edges, curr_map = set(),  dict()

	with open(ext_file) as f:
		for line in f:
			if line.startswith('regulator_name'): continue
			line = line.strip()
			s = re.split(',',line)
			name1, name2 = getVariableName(model_dict,curr_map,s[0:regulated_col]), getVariableName(model_dict,curr_map,s[regulated_col:interaction_col])
			if name1=="" or name2=="": continue
			pos = '+' if s[16]=='increases' else '-'

			if (name2 in model_dict and name1 in model_dict[name2]):
				continue
			ext_edges.add((name1, name2, pos))
	return ext_edges

def merge_clusters(regulators, path, ReturnTh):
    """
    This function records indices of clusters to be merged based on the existence of return paths.
    It generates the grouped_ext_Merged pickle file that contains the merged clusters.

    Parameters
    ----------
    regulators : dict
		Contains baseline model elements and corresponding regulator elements
    path : str
		The path of the directory that contains the grouped_ext file
    ReturnTh : int
		A user-defined integer threshold for the number of return paths, beyond which clusters will be merged
    """
    # Merge clusters if there is one or more return paths

    G = nx.DiGraph()
    G = makeDiGraphBase(regulators)
    com_edges = list()
    group_num = 1
    extensions = pickle.load(open(os.path.join(path,"grouped_ext"),'rb'))

    for ii in range(0,len(extensions)):
        for jj in range(ii+1,len(extensions)):
            count = 0
            cluster1 = extensions[ii]
            cluster2 = extensions[jj]
            G1 = nx.DiGraph()
            G2 = nx.DiGraph()

            for e in cluster1[1:]:
                #ee=e[1].split('->')
                if e[2] == '+':
                    G1.add_edge(e[0], e[1],weight=1)
                elif e[2] == '-':
                    G1.add_edge(e[0], e[1],weight=0)

            for e in cluster2[1:]:
                #ee=e[1].split('->')
                if e[2] == '+':
                    G2.add_edge(e[0], e[1],weight=1)
                elif e[2] == '-':
                    G2.add_edge(e[0], e[1],weight=0)
            Gall = nx.compose(G1,G2)
            for g in G.edges:
                if g[0] in G1:
                    for ne in G.successors(g[1]):
                        if ne in G2:
                            for ne1 in G2.successors(ne):
                                if ne1 in G:
                                    count = count+1
                if g[0] in G1:
                    for ne in G.successors(g[0]):
                        if ne in G2:
                            for ne1 in G2.successors(ne):
                                if ne1 in G:
                                    count = count+1
            if count > int(ReturnTh): #set threshold for the number of return paths
                logging.info('Merge clusters NO.{} and NO.{}'.format(str(ii+1),str(jj+1)))
                #print(str(ii) + " and " + str(jj))
                Gall = nx.compose(G1,G2)
                NODESS = list()
                for (node1,node2,data) in Gall.edges(data=True):
                    temp=list()
                    temp.append(node1)
                    temp.append(node2)
                    if data['weight'] == 0:
                        temp.append('-')
                    elif data['weight'] == 1:
                        temp.append('+')
                    NODESS.append(temp)
                com_edges.append([group_num] + NODESS)
                group_num = group_num+1

    pickle.dump(com_edges, open(os.path.join(path, "grouped_ext_Merged"),'wb')) #Merged clusters

    return


def makeDiGraphBase(regulators):
    """
    A utility function for merge_clusters(), this function converts the baseline model into a directed graph.

    Parameters
    ----------
    regulators : dict
		Contains baseline model elements and corresponding regulator elements

    Returns
    -------
    G : DiGraph()
        Directed graph of the baseline model
    """
    G = nx.DiGraph()
    G.clear()
    for key, values in regulators.items():
        G.add_node(key)
        for value in values:
            G.add_edge(value, key)
    return G

def getRow(mdl):
	"""
 	A utility function for extend_model(), this function returns a dict indicating the row of each element of the parsed model.

	Parameters
	----------
	mdl : str
		The path that will be used to store new extended model spreadsheet

	Returns
	-------
	res : dict
        A dict indicating the row number of each element
	"""

	df = pd.read_excel(mdl)
	res = dict()
	for i in df.index:
		el=df['Element name'][i].strip()
		res[el]=i+2
	return res

def extend_model(base_mdl,clusters,ext_mdl):
	"""
 	This function creates a new xlsx for each candidate model (i.e. baseline model + cluster)

	Parameters
	----------
	base_mdl : str
		The path that contains the baseline model spreadsheet to be extended
	clusters : list
		It holds the edges inside each grouped extension
	ext_mdl :	 str
		The path that will be used to store new extended model spreadsheet
	"""

	os.system('cp '+base_mdl+' '+ext_mdl)
	df = pd.read_excel(ext_mdl)
	pos=df.columns.get_loc("Positive regulators")+1
	neg=df.columns.get_loc("Negative regulators")+1
	var_name=df.columns.get_loc("Variable name")+1
	ini0=df.columns.get_loc("Initial 0")+1 #uncomment the following lines if you have more than one scenario in the properties
	ini1=df.columns.get_loc("Initial 1")+1
	ini2=df.columns.get_loc("Initial 2")+1
#	ini3=df.columns.get_loc("Initial 3")+1
#	ini4=df.columns.get_loc("Initial 4")+1
#	ini5=df.columns.get_loc("Initial 5")+1
	el_name=df.columns.get_loc("Element name")+1
	name_to_row = getRow(ext_mdl)
	curr_row = len(name_to_row)+2
	wb = openpyxl.load_workbook(ext_mdl)
	ws = wb.active
	for e in clusters[1:]:
		if e[0] not in name_to_row:
			ws.cell(row=curr_row,column=el_name,value=e[0])
			ws.cell(row=curr_row,column=var_name,value=e[0])
			ws.cell(row=curr_row,column=ini0,value=1) #uncomment the following lines if you have more than one scenario in the properties
			ws.cell(row=curr_row,column=ini1,value=1)
			ws.cell(row=curr_row,column=ini2,value=1)
#			ws.cell(row=curr_row,column=ini3,value=1)
#			ws.cell(row=curr_row,column=ini4,value=1)
#			ws.cell(row=curr_row,column=ini5,value=1)
			name_to_row[e[0]] = curr_row
			curr_row += 1
		if e[1] not in name_to_row:
			ws.cell(row=curr_row,column=el_name,value=e[1])
			ws.cell(row=curr_row,column=var_name,value=e[1])
			ws.cell(row=curr_row,column=ini0,value=1) #uncomment the following lines if you have more than one scenario in the properties
			ws.cell(row=curr_row,column=ini1,value=1)
			ws.cell(row=curr_row,column=ini2,value=1)
#			ws.cell(row=curr_row,column=ini3,value=1)
#			ws.cell(row=curr_row,column=ini4,value=1)
#			ws.cell(row=curr_row,column=ini5,value=1)
			name_to_row[e[1]] = curr_row
			curr_row += 1
		col = pos if e[2]=='+' else neg
		original = ws.cell(row=name_to_row[e[1]],column=col).value
		original = original+',' if original!=None else ''
		ws.cell(row=name_to_row[e[1]],column=col,value=original+e[0])
	wb.save(ext_mdl)

def get_args():

    parser = argparse.ArgumentParser(description="Network model extension using ACCORDION")
    parser.add_argument('ReadingOutput', type=str,help="Reading output spreadsheet")
    parser.add_argument('Baseline', type=str,help="Baseline model in BioRECIPES format")
    parser.add_argument('Inflation', type=str,help="Inflation parameter for Markov clustering")
    parser.add_argument('ReturnTh', type=str,help="Return path threshold")
    parser.add_argument('out', type=str,help="Output directory")
    args = parser.parse_args()
    return(args)

def main():

    args = get_args()

    t0 = time.time()
    #Reading output .csv file. File format(RegulatedName,RegulatedID,RegulatedType,RegulatorName,RegulatorID,RegulatorType,PaperID)
    interaction_filename = args.ReadingOutput

    #Baseline model
    model_dict, regulators = get_model(args.Baseline)

    #use parseExtension if Reading output format is (RegulatedName,RegulatedID,RegulatedType,RegulatorName,RegulatorID,RegulatorType,PaperID)
    exttt=parseExtension(model_dict, interaction_filename)

    res,new_base_model = runMarkovCluster(args.out,exttt,model_dict,args.Inflation) # try 1.5,2,4,6

    merge_clusters(regulators, args.out, args.ReturnTh)

	# for example, extend the baseline model to include the first cluster from the unmerged result 'res' and generate 'extension.xlsx'
	#extend_model(args.Baseline,res[0],args.out+'extension.xlsx')

    t1 = time.time()
    total = t1-t0
    logging.info("Time to run ACCORDION in seconds: {}".format(str(total)))

if __name__ == '__main__':
    main()
