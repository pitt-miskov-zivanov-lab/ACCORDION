import re, copy, sys, os
import openpyxl
from collections import defaultdict
import pandas as pd
import pickle

def runMarkovCluster(out_dir,ext_edges,base_model,coef):
	"""
 	This function prepares the inputs to the markov clustering algorithm and  
 	creates a pickle file for the output clusters. It also creates a file for
 	each cluster.
   
	Parameters
	----------
 	out_dir : str
        Path of the directory that will include the output clusters files
 	ext_edges : set
        Holds the interactions in the reading output file. Each interaction is 
        in the form
        (regulator element, regulated element, type of interaction (+/-))
 	base_model : dict
        Dictionary that holds baseline model regulator and regulated elements
 	coef : int
        The inflation parameter of the markov clustering algorithm
          
	Returns
	-------
  	res : list
        Each candidate in this list is a cluster 
        new_base_model : dict
        The baseline model elements are keys of this dict and the values are the corresponindg regulator elements
	"""
 
	ext_model,new_base_model = buildExtGraph(ext_edges,base_model)
	clusteringAlgo(out_dir,ext_model,coef)
	ModelNetwork(out_dir,base_model)
	res = getGroupedExt(out_dir+'markov_cluster',ext_edges)
	pickle.dump(res, open(out_dir+'grouped_ext','wb'))
	return res,new_base_model

def buildExtGraph(ext_edges,base_model=dict()):
	"""
 	This function constructs the graph for the extensions and the baseline model
	
	Parameters
	----------
	ext_edges : set
        Holds the interactions in the reading output file. Each interaction is 
        in the form
        (regulator element, regulated element, type of interaction (+/-))
 	base_model : dict
        Dictionary that holds baseline model regulator and regulated elements
        
	Returns
	-------
	ext_model : dict
    	  This dict contains the elements of both the baseline model and the reading output file. 
	  Those elements are the keys and the values are the corresponding regulator elements. 
        new_base_model : dict
        The baseline model elements are keys of this dict and the values are the corresponindg regulator elements
	"""
    
	new_base_model=dict()
	for k in base_model:
		new_base_model[k]=base_model[k]['regulators']        
	ext_model = copy.deepcopy(new_base_model)
	for edge in ext_edges:
		if edge[1] in ext_model:
			ext_model[edge[1]].add(edge[0])          
		else:
			ext_model[edge[1]] = {edge[0]}
	return ext_model,new_base_model

def clusteringAlgo(out_dir,ext_model,coef):
	"""
 	This function is designed to run MCL
	
	Parameters
	----------
	out_dir : str
        Path of the directory that will include the output clusters files
	ext_model : dict
        This dict contains the elements of both the baseline model and the reading output file. 
        Those elements are the keys and the values are the corresponding regulator elements. 
	coef : int
        The inflation parameter of the markov clustering algorithm
	Returns
	-------
  	res : list
        Each candidate in this list is a cluster 
	"""
	
	cluster_file = out_dir + 'markov_cluster'
	abc_model = out_dir + 'abc_model'
	output_stream = open(abc_model, 'w')
	for tgt in sorted(ext_model):
		for reg in sorted(ext_model[tgt]):
			if tgt==reg: continue
			output_stream.write(reg+' '+tgt+' '+str(1)+'\n')
	output_stream.close()
	print('mcl '+abc_model.replace(' ','\ ')+' --abc -I '+str(coef)+' -o '+cluster_file.replace(' ','\ '))
	os.system('mcl '+abc_model.replace(' ','\ ')+' --abc -I '+str(coef)+' -o '+cluster_file.replace(' ','\ '))

def ModelNetwork(out_dir,base_model):
	"""
 	This function creates the network of baseline model and machine reading output
	
	Parameters
	----------
	out_dir : str
        Path of the directory that will include the output clusters files
 	base_model : dict
        Dictionary that holds baseline model regulator and regulated elements
	Returns
	-------
	"""
    
	new_base_model=dict()
	for k in base_model:
		new_base_model[k]=base_model[k]['regulators'] 	
	abc_model = out_dir + 'abc_model_network'
	output_stream = open(abc_model, 'w')
	for tgt in sorted(new_base_model):
		for reg in sorted(new_base_model[tgt]):
			if tgt==reg: continue
			output_stream.write(reg+' '+tgt+'\n')
	output_stream.close()

def getGroupedExt(cluster_file,ext_edges):
	"""
 	This function generates list of extensions
	
	Parameters
	----------
	cluster_file : str
	  The path of the markov_cluster file
	ext_edges : set
        Holds the interactions in the reading output file. Each interaction is 
        in the form
        (regulator element, regulated element, type of interaction (+/-))
        
	Returns
	-------
  	res : list
        Each candidate in this list is a cluster 
	"""
 
	group_num = 1
	get_group = dict()
	get_ext = defaultdict(list)
	res = list()
	with open(cluster_file) as f:
		for idx, line in enumerate(f,start=1):
			for ele in re.findall('\S+',line.strip()):
				#print(idx)
				get_group[ele] = idx

	for edge in ext_edges:
		g1 = get_group[edge[0]] if edge[0] in get_group else sys.maxsize
		g2 = get_group[edge[1]] if edge[1] in get_group else sys.maxsize
		group = min(g1, g2)
		if group==sys.maxsize: continue
		get_ext[group] += [list(edge)]

	for key in sorted(get_ext):
		if not get_ext[key]: continue
		res.append([group_num]+get_ext[key])
		group_num += 1

	return res

def getRow(mdl):
	"""
 	This function returns a dict indicating the row of each element of the parsed model.
	
	Parameters
	----------
	mdl : dict
        This dict contains the elements of either the baseline model or candidate model. 
        Those elements are the keys and the values are the corresponding regulator elements. 

	Returns
	-------
	res : dict
        A dict indicating the row of each element
	"""
    
	df = pd.read_excel(mdl)
	res = dict()
	for i in df.index:
		el=df['Element name'][i].strip()
		res[el]=i+2
	return res
 
def extend_model(mdl,edges,ext_model):
	"""
 	This function opens a new xlsx for each candidate model (i.e. baseline model + cluster) 
	
	Parameters
	----------
	mdl : dict
        Dictionary that holds baseline model regulator and regulated elements
	edges : list
	  It holds the edges of each parsed cluster
	ext_model :	 dict
    	  This dict contains the elements of both the baseline model and the reading output file. 
	  Those elements are the keys and the values are the corresponding regulator elements. 

	Returns
	-------
	"""
 
	os.system('cp '+mdl+' '+ext_model)
	df = pd.read_excel(ext_model)
	pos=df.columns.get_loc("Positive regulators")+1
	neg=df.columns.get_loc("Negative regulators")+1
	var_name=df.columns.get_loc("Variable name")+1
	ini0=df.columns.get_loc("Initial 0")+1 #uncomment the following lines if you have more than one scenario in the properties
#	ini1=df.columns.get_loc("Initial 1") +1
#	ini2=df.columns.get_loc("Initial 2") +1
#	ini3=df.columns.get_loc("Initial 3")+1
#	ini4=df.columns.get_loc("Initial 4") +1
#	ini5=df.columns.get_loc("Initial 5") +1
	el_name=df.columns.get_loc("Element name")   +1
	name_to_row = getRow(ext_model)
	curr_row = len(name_to_row)+2
	wb = openpyxl.load_workbook(ext_model)
	ws = wb.active
	for e in edges[1:]:
		if e[0] not in name_to_row: 
			ws.cell(row=curr_row,column=el_name,value=e[0])
			ws.cell(row=curr_row,column=var_name,value=e[0])
			ws.cell(row=curr_row,column=ini0,value=1) #uncomment the following lines if you have more than one scenario in the properties
#			ws.cell(row=curr_row,column=ini1,value=1)
#			ws.cell(row=curr_row,column=ini2,value=1)
#			ws.cell(row=curr_row,column=ini3,value=1)
#			ws.cell(row=curr_row,column=ini4,value=1)
#			ws.cell(row=curr_row,column=ini5,value=1)
			name_to_row[e[0]] = curr_row
			curr_row += 1
		if e[1] not in name_to_row:
			ws.cell(row=curr_row,column=el_name,value=e[1])
			ws.cell(row=curr_row,column=var_name,value=e[1])
			#print(e[1])           
			ws.cell(row=curr_row,column=ini0,value=1)
#			ws.cell(row=curr_row,column=ini1,value=1)
#			ws.cell(row=curr_row,column=ini2,value=1)
#			ws.cell(row=curr_row,column=ini3,value=1)
#			ws.cell(row=curr_row,column=ini4,value=1)
#			ws.cell(row=curr_row,column=ini5,value=1)
			name_to_row[e[1]] = curr_row
			curr_row += 1
		col = pos if e[2]=='+' else neg
		original = ws.cell(row=name_to_row[e[1]],column=col).value
		original = original+',' if original!=None else ''
		ws.cell(row=name_to_row[e[1]],column=col,value=original+e[0])
	wb.save(ext_model)


